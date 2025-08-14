import io
from typing import Sequence, Tuple
from urllib.parse import unquote

import openpyxl

from backend.api.entities.category import Category
from backend.api.entities.general import General
from backend.api.repositories.category import CategoryRepo
from backend.api.repositories.general import GeneralRepo
from backend.api.std import func


def get_general(p_code: int, p_category: str) -> General:
    """
    汎用マスタ情報取得
        Args:
            p_code: 汎用マスタコード
            p_category: 値
        Returns:
            general: 取得した汎用マスタ情報
    """

    # 汎用マスタ管理クラスのインスタンス生成
    if p_code == -1:
        return General()
    else:
        general_repo = GeneralRepo()
        return general_repo.find_by_code(p_code, p_category)


def get_select_list() -> Sequence[Category]:
    """
    汎用マスタ情報一覧取得処理
    Args:
    Returns:
        Sequence: 検索結果のリスト
    """
    category_repo = CategoryRepo()
    return category_repo.list()


def search_general_list(
    sel_category: str, txt_code_value: str, offset: int, limit: int
) -> Tuple[Sequence[General], int]:
    """
    汎用マスタ一覧取得処理
    Args:
        sel_category: カテゴリ（完全一致）
        txt_code_value: 名称（部分一致）
        offset: 何件目から取得するか指定（全件表示する場合は-1を指定）
        limit: 何件取得するか指定
    Returns:
        list: 検索結果のリスト
        int: レコード件数

    """
    general_repo = GeneralRepo()
    return general_repo.find_list(sel_category, txt_code_value, offset, limit)


def create_general(general: General) -> int:
    """
    汎用マスタ情報 新規登録
        Args:
            general: 登録する汎用マスタ情報
        Returns:
            int: 登録したレコードのID
    """

    general_repo = GeneralRepo()
    return general_repo.create(general)


def update_general(general: General) -> General:
    """
    機種情報 更新
        Args:
            general: 更新する汎用マスタ情報
        Returns:
            general: 取得した汎用マスタ情報

    """

    general_repo = GeneralRepo()
    general_repo.update(general)
    return general_repo.find_by_code(general.code, general.category)


def create_download_file(general_list: Sequence[General]) -> io.BytesIO:
    """
    ダウンロードファイル作成
    Args:
        general_list: 検索結果リスト
    Returns:
        io.BytesIO: 生成したExcelファイル（file-likeオブジェクト）

    """

    # Excelデータ出力内容定義用リスト作成
    header_lst = [
        "カテゴリ",
        "コード",
        "名称",
        "備考",
        "ソートキー",
        "最終更新者",
        "最終更新日時",
        "削除フラグ",
    ]

    xl = io.BytesIO()
    wb = openpyxl.Workbook()
    sht = wb.worksheets[0]

    # ヘッダー行作成
    for c_idx in range(len(header_lst)):
        sht.cell(row=1, column=c_idx + 1, value=header_lst[c_idx])

    # 2行目からデータ部
    r_idx = 2
    for general in general_list:
        sht.cell(row=r_idx, column=1, value=general.r_category.display_name)
        sht.cell(row=r_idx, column=2, value=str(general.code))
        sht.cell(row=r_idx, column=3, value=general.code_value)
        sht.cell(row=r_idx, column=4, value=general.remarks)
        sht.cell(row=r_idx, column=5, value=str(general.sort_key))
        sht.cell(row=r_idx, column=6, value=general.update_user)
        sht.cell(
            row=r_idx, column=7, value=general.update_date.strftime("%Y/%m/%d %H:%M:%S")
        )
        sht.cell(row=r_idx, column=8, value=str(general.del_flag))
        r_idx += 1

    # 書式設定(罫線と背景色)
    func.set_sheets_orgformat(sht, True)

    # Excel保存
    wb.save(xl)
    xl.seek(0)

    return xl


def set_search_decode(sel_category: str, txt_code_value: str) -> Tuple[str, str]:
    """
    検索条件に対してURLデコード
    Args:
        sel_category, txt_code_value: 検索キーワード
    Returns:
        str: sel_category（デコード後）
        str: txt_code_value（デコード後）
    """
    return unquote(sel_category), unquote(txt_code_value)


def duplicate_check(general: General) -> bool:
    """
    同カテゴリ内の同一の名称の有無のチェック
    Args:
        重複チェック対象の汎用関連情報
    Returns:
        bool: m_general内の重複データ有無の結果（重複データが0件の場合はtrue）
    """
    general_repo = GeneralRepo()
    return general_repo.duplicate_check(general)
