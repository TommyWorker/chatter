import io
from typing import Dict, Optional, Sequence, Tuple
from urllib.parse import unquote

import openpyxl

from backend.api.entities.general import General
from backend.api.entities.user import User
from backend.api.repositories.general import GeneralRepo
from backend.api.repositories.user import UserRepo
from backend.api.std import func


def get_user(id: int) -> User:
    """
    ユーザ情報取得
        Args:
            id: ユーザID
        Returns:
            User: 取得したユーザ情報
    """

    # ユーザクラスのインスタンス生成
    if id == 0:
        return User()
    else:
        user_repo = UserRepo()
        return user_repo.find_by_id(id)


def create_user(user: User) -> User:
    """
    ユーザ情報 新規登録
        Args:
            user: 登録するユーザ情報
        Returns:
            User: 登録したユーザ情報
    """

    user_repo = UserRepo()
    user_id = user_repo.create(user)
    return user_repo.find_by_id(user_id)


def update_user(user: User) -> User:
    """
    ユーザ情報 更新
        Args:
            user: 更新するユーザ情報
        Returns:
            user: 更新したユーザ情報

    """

    user_repo = UserRepo()
    user_repo.update(user)
    assert user.id is not None
    return user_repo.find_by_id(user.id)


def find_users( 
    mail_address: Optional[str] = None,
    user_name: Optional[str] = None,
    offset: int = 0,
    limit: int = 10,
) -> Tuple[Sequence[User], int]:
    """
    ユーザ情報一覧取得処理
    Args:
        mail_address: メールアドレス（部分一致）
        user_name: 名前（部分一致）
        offset: 何件目から取得するか指定（全件表示する場合は-1を指定）
        limit: 何件取得するか指定
    Returns:
        list: 検索結果のリスト
        int: レコード件数
    """
    user_repo = UserRepo()
    return user_repo.find(mail_address, user_name, offset, limit)


def create_download_file(user_list: Sequence[User]) -> io.BytesIO:
    """
    ダウンロードファイル作成
    Args:
        user_list: 検索結果リスト
    Returns:
        io.BytesIO: 生成したExcelファイル（file-likeオブジェクト）

    """

    # Excelデータ出力内容定義用リスト作成
    header_lst = [
        "ID",
        "メールアドレス",
        "名前",
        "権限",
        "最終更新者",
        "最終更新日時",
        "削除フラグ",
    ]

    xl = io.BytesIO()
    wb = openpyxl.Workbook()
    sht = wb.worksheets[0]

    # ヘッダー行作成
    for c_idx in range(len(header_lst)):
        sht.cell(row=1, column=c_idx + 1, value=header_lst[c_idx])

    # 2行目からデータ部
    r_idx = 2
    for user in user_list:
        c_idx = 1
        sht.cell(row=r_idx, column=c_idx, value=str(user.id))
        c_idx += 1
        sht.cell(row=r_idx, column=c_idx, value=user.mail_address)
        c_idx += 1
        sht.cell(row=r_idx, column=c_idx, value=user.user_name)
        c_idx += 1
        sht.cell(row=r_idx, column=c_idx, value=user.authority[0].code_value)
        c_idx += 1
        sht.cell(row=r_idx, column=c_idx, value=user.update_user)
        c_idx += 1
        sht.cell(
            row=r_idx,
            column=c_idx,
            value=user.update_date.strftime("%Y/%m/%d %H:%M:%S"),
        )
        c_idx += 1
        sht.cell(row=r_idx, column=c_idx, value=str(user.del_flag))
        r_idx += 1

    # 書式設定(罫線と背景色)
    func.set_sheets_orgformat(sht, True)

    # Excel保存
    wb.save(xl)
    xl.seek(0)

    return xl


def set_search_decode(user_name: str) -> str:
    """
    検索条件に対してURLデコード
    Args:
        user_name: 検索キーワード
    Returns:
        str: txt_user（デコード後）
    """
    return unquote(user_name)


def check_user(user: User, mode: str) -> Tuple[bool, str]:
    """
    ユーザ情報更新時の入力チェック
    Args:
        user: 処理対象のユーザ情報
        mode: new or update
    Returns:
        Tuple1: 完了判定（True：正常 False：エラー）
        Tuple2: エラーメッセージ
    """

    # 新規処理の場合、メールアドレスの存在チェック
    if mode == "new":
        user_repo = UserRepo()
        dmy_user, count = user_repo.find(user.mail_address)
        if count > 0:
            return False, "入力したメールアドレスは既に存在します"

    return True, ""


def get_selected_lists() -> Dict[str, Sequence[General]]:
    """
    画面の選択項目のリスト取得
        Args:

        Returns:
            以下のリストを返す辞書
                authority: 権限リスト
    """

    general_repo = GeneralRepo()

    authority_list = general_repo.find("authority_code")

    return {
        "authority": authority_list,
    }
